#!/usr/bin/env python3
"""Import FANS PICK or SHOW! MUSIC CORE winners from .xlsx/.csv.

The file must contain email and nickname columns. Music Core also requires an
event date, either in the file or through --event-date. Writes require --apply.
Secrets are read from SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Iterable

EMAIL_ALIASES = {"email", "account_email", "가입 이메일", "이메일", "muniverse 가입 이메일", "muniverse 이메일"}
NICK_ALIASES = {"nickname", "muniverse_nickname", "닉네임", "muniverse 닉네임"}
DATE_ALIASES = {"event_date", "recording_date", "녹화일", "대상 녹화일"}


def canonical(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def normalize_email(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "").strip()).lower()


def normalize_nickname(value: Any, kind: str) -> str:
    text = unicodedata.normalize("NFKC", str(value or "").strip())
    text = " ".join(text.split())
    return text.lower() if kind == "fans-pick" else text


def sha256(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def rows_from_xlsx(path: Path) -> Iterable[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required: pip install -r tools/requirements.txt") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [canonical(value) for value in next(iterator)]
    except StopIteration:
        return
    for values in iterator:
        yield {headers[index]: value for index, value in enumerate(values) if index < len(headers)}


def rows_from_csv(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            yield {canonical(key): value for key, value in row.items()}


def find_value(row: dict[str, Any], aliases: set[str]) -> Any:
    canonical_aliases = {canonical(alias) for alias in aliases}
    for key, value in row.items():
        if canonical(key) in canonical_aliases:
            return value
    return ""


def request_json(url: str, key: str, method: str = "GET", body: Any | None = None) -> Any:
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(url, data=data, method=method)
    request.add_header("apikey", key)
    request.add_header("Authorization", f"Bearer {key}")
    request.add_header("Content-Type", "application/json")
    request.add_header("Prefer", "return=minimal")
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            payload = response.read().decode("utf-8")
            return json.loads(payload) if payload else None
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Supabase HTTP {exc.code}: {details[:500]}") from exc


def existing(base: str, key: str, table: str, identity_hash: str, event_date: str | None) -> bool:
    params = {"select": "id", "identity_hash": f"eq.{identity_hash}", "limit": "1"}
    if event_date:
        params["event_date"] = f"eq.{event_date}"
    url = f"{base}/rest/v1/{table}?{urllib.parse.urlencode(params)}"
    rows = request_json(url, key)
    return isinstance(rows, list) and bool(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("file", type=Path)
    parser.add_argument("--kind", choices=("fans-pick", "music-core"), required=True)
    parser.add_argument("--event-date", help="YYYY-MM-DD; required for Music Core when absent in the file")
    parser.add_argument("--apply", action="store_true", help="write to Supabase; otherwise validate only")
    args = parser.parse_args()

    if not args.file.exists():
        raise SystemExit(f"File not found: {args.file}")
    source_rows = rows_from_xlsx(args.file) if args.file.suffix.lower() in {".xlsx", ".xlsm"} else rows_from_csv(args.file)

    records: list[dict[str, Any]] = []
    seen: set[tuple[str, str | None]] = set()
    errors: list[str] = []
    for number, row in enumerate(source_rows, start=2):
        email = normalize_email(find_value(row, EMAIL_ALIASES))
        nickname = normalize_nickname(find_value(row, NICK_ALIASES), args.kind)
        event_date = str(find_value(row, DATE_ALIASES) or args.event_date or "").strip() or None
        if not email or "@" not in email or not nickname:
            errors.append(f"row {number}: email or nickname is missing/invalid")
            continue
        if args.kind == "music-core" and (not event_date or len(event_date) != 10):
            errors.append(f"row {number}: event date is required")
            continue
        identity_hash = sha256(f"{email}\n{nickname}")
        unique = (identity_hash, event_date)
        if unique in seen:
            errors.append(f"row {number}: duplicate winner")
            continue
        seen.add(unique)
        if args.kind == "fans-pick":
            records.append({"identity_hash": identity_hash, "submitted": False})
        else:
            records.append({
                "identity_hash": identity_hash,
                "email_hash": sha256(email),
                "nickname_hash": sha256(nickname),
                "event_date": event_date,
                "submitted": False,
            })

    if errors:
        print("Validation warnings:", file=sys.stderr)
        for error in errors:
            print(f"- {error}", file=sys.stderr)
    if not records:
        raise SystemExit("No valid winner records found")
    print(f"Validated {len(records)} winner record(s).")
    if not args.apply:
        print("Dry run only. Re-run with --apply to write.")
        return 0

    base = os.environ.get("SUPABASE_URL", "").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not base or not key:
        raise SystemExit("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required")
    table = "cover_pick_winners" if args.kind == "fans-pick" else "music_core_winners"
    inserted = skipped = 0
    for record in records:
        event_date = record.get("event_date")
        if existing(base, key, table, record["identity_hash"], event_date):
            skipped += 1
            continue
        request_json(f"{base}/rest/v1/{table}", key, method="POST", body=record)
        inserted += 1
    print(f"Inserted {inserted}; already present {skipped}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
